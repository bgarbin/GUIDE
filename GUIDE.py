# -*- coding: utf-8 -*-

import pyqtgraph as pg
from pyqtgraph.Qt import QtCore, QtGui
from pyqtgraph.dockarea import *
from pyqtgraph.ptime import time as pgtime
from pyqtgraph.graphicsItems.GradientEditorItem import Gradients

import sys
import numpy as np
import os
from functools import partial
import time
from openpyxl import load_workbook
import pandas as pd
import time

pg.mkQApp()

## Define main window class from template
path = os.path.dirname(os.path.abspath(__file__))
uiFile = os.path.join(path, 'GUIDE.ui')
WindowTemplate, TemplateBaseClass = pg.Qt.loadUiType(uiFile)


### BEGIN Modele class ###
class Modele():

    def __init__(self):

        # Allow importing any file provided as argument in the form: python3 GUIDE.py -f model_input
        if len(sys.argv) > 1:
            import importlib
            option_name = sys.argv[1]
            assert option_name == '-f', f"option '{option_name}' not understood. Known option is only '-f' with filename as value"
            assert len(sys.argv)>=3, "provide a filename to load parameters from"
            lib_path = sys.argv[2]
            if lib_path.endswith('.py'): lib_path = lib_path.rstrip('.py')
            input_file = importlib.import_module(lib_path.replace('/','.'))
        else:
            import model_input as input_file

        # Loading plots configuration (used in MainWindow class)
        docks = input_file.load_docks()
        setattr(self,'docks',docks)

        # Loading window parameters
        window_params = input_file.window_params
        if 'streaming' not in window_params.keys(): self.streaming = True
        for window_param in window_params.keys():
            setattr(self,window_param,window_params[window_param])

        # Tracking time
        self.nstep      = 0
        self.time_stamp = np.zeros(self.array_size).astype(np.float64)

        # Loading parameters
        params = input_file.load_params()
        setattr(self,'params',params)
        for param in self.params.keys():
            if isinstance(self.params[param]['step'],int):  typ = int
            else: typ = np.float64
            self.params[param]['value'] = self.params[param]['init_cond'] * np.ones(self.array_size).astype(typ)

        # Loading variables
        variables = input_file.load_variables()
        setattr(self,'variables',variables)
        # Loading observables
        observables = input_file.load_observables()
        setattr(self,'observables',observables)

        # List as defined in the input file (as observables are added to variables dict)
        list_variables = list(self.variables.keys())
        list_observables = list(self.observables.keys())

        # Concatenate the variables and observables dict (if 'invert_var_obs' then invert order observables and variables are displayed)
        if not 'invert_order_obs_var' in window_params.keys(): self.variables = dict(self.variables, **self.observables)
        else:
            if window_params['invert_order_obs_var']:
                self.variables = dict(self.observables, **self.variables)
            else:
                self.variables = dict(self.variables, **self.observables)

        # Build main dict of variables
        for variable in self.variables.keys():
            self.variables[variable]['value'] = self.variables[variable]['init_cond'] * np.ones(self.array_size).astype(self.variables[variable]['type'])
            if variable in list_variables:
                self.variables[variable]['observable'] = False
            elif variable in list_observables:
                self.variables[variable]['observable'] = True

        # Set default plot to True if none provided
        for variable in self.variables.keys():
            if 'plot' not in self.variables[variable].keys():
                self.variables[variable]['plot'] = True

        # Loading equations into keyword 'equation' in variables dict
        # 'diff_eq_' and 'eq_' are default patterns for variables and observables respectively
        pattern_variables = 'diff_eq_'
        for key in [attr for attr in input_file.__dict__.keys() if attr.startswith(pattern_variables)]:
            variable = key.split(pattern_variables)[-1]
            assert variable in list_variables, f"Variable {variable} not understood or badly declared"
            if 'equation' in self.variables[variable].keys(): continue
            self.variables[variable]['equation'] = input_file.__dict__[key]

        pattern_observables = 'eq_'
        for key in [attr for attr in input_file.__dict__.keys() if attr.startswith(pattern_observables)]:
            variable = key.split(pattern_observables)[-1]
            assert variable in list_observables, f"Observable {variable} not understood or badly declared"
            if 'equation' in self.variables[variable].keys(): continue
            self.variables[variable]['equation'] = input_file.__dict__[key]

        # Create dict of the usable kernels
        self.kernels = {}
        pattern_kernels = 'kernel_'
        for key in [attr for attr in self.__dir__() if attr.startswith(pattern_kernels)]:
            kernel = key.split(pattern_kernels)[-1]
            self.kernels[kernel] = {}
            self.kernels[kernel]['value'] = getattr(self,key)
        for key in [attr for attr in input_file.__dict__.keys() if attr.startswith(pattern_kernels)]:
            kernel = key.split(pattern_kernels)[-1]
            self.kernels[kernel] = {}
            self.kernels[kernel]['value'] = input_file.__dict__[key]

        # Load additional keyboard keys if any provided
        self.user_defined_keyPressEvent = input_file.keyboard_keys()
        if self.user_defined_keyPressEvent is None: self.user_defined_keyPressEvent = {} # if None provided
        system_reserved_keys = [" ", "q", "h", "s", "r", "i", "c"]

        for user_defined_key in self.user_defined_keyPressEvent.keys():
            assert user_defined_key not in system_reserved_keys, f"User defined key '{user_defined_key}' in system reserved ones {system_reserved_keys}"


        ########################### BEGIN Assertions input file ###########################
        # 'dock' (variables): Not providing dock_name that doesn't exist
        for variable in self.variables.keys():
            if 'dock' in self.variables[variable]:
                for dock_name in self.variables[variable]['dock']:
                    if not isinstance(dock_name,dict):
                        assert dock_name in self.docks.keys(), f"Dock name '{dock_name}' for variable {variable} not understood. Dock name must be in {list(self.docks.keys())}"
        # all variables have an equation
        for variable in self.variables.keys():
            assert 'equation' in self.variables[variable].keys(), f"An equation for variable {variable} must be provided"
        ###########################  END Assertions input file  ###########################


    def simulator(self):

        """ Calculate 1 time step and update arrays """

        # Actual computation (pass only the 'value' keyword of each sub-dictionnary)
        self.computation_result_dict = self.kernels[self.kernel]['value']({key:value['value'][-1] for (key,value) in self.variables.items() if not value['observable']},{key:value['value'][-1] for (key,value) in self.params.items()})  # use last value of all variables for the computations of next step

        # Update last values to the newest calculated
        for variable in self.variables.keys():
            if not self.variables[variable]['observable']:
                # Simpler concatenate replacing directly indices
                self.variables[variable]['value'][:-1] = self.variables[variable]['value'][1:]
                self.variables[variable]['value'][-1] = self.computation_result_dict[variable]

        # Evaluate observables
        self.update_observables()

    def update_observables(self):
        for variable in self.variables.keys():
            if self.variables[variable]['observable']:
                self.obs_computation_result = self.variables[variable]['equation'](self,{key:value['value'] for (key,value) in self.variables.items()},{key:value['value'][-1] for (key,value) in self.params.items()})
                if 'calculation_size' in self.variables[variable].keys() and self.variables[variable]['calculation_size']:
                    self.variables[variable]['value'] = self.obs_computation_result
                else:
                    try:              index = len(self.obs_computation_result)
                    except TypeError: index = 1   # If return only a single value
                self.variables[variable]['value'][:-index] = self.variables[variable]['value'][index:]
                self.variables[variable]['value'][-index:] = self.obs_computation_result


    def kernel_euler(self, variables, params):

        """ N variables Euler algorithm (A = A + dt * eq_A(params)) """

        new_variables = {}
        for variable_name in variables.keys():
            new_variables[variable_name] = variables[variable_name] + self.step_size * self.variables[variable_name]['equation'](self,variables,params)

        return new_variables

    def kernel_RK4(self, variables, params):

        """ N variables RK4 algorithm """

        temp_variables = variables.copy()

        # Loop for each coefficient on all equations
        coefs_1 = {}
        for variable_name in variables.keys():
            coefs_1[variable_name] = self.variables[variable_name]['equation'](self,temp_variables,params)

        coefs_2 = {}
        for variable_name in variables.keys():    # evaluate variables first
            temp_variables[variable_name] = variables[variable_name] + (self.step_size/2.)*coefs_1[variable_name]
        for variable_name in variables.keys():
            coefs_2[variable_name] = self.variables[variable_name]['equation'](self,temp_variables,params)

        coefs_3 = {}
        for variable_name in variables.keys():
            temp_variables[variable_name] = variables[variable_name] + (self.step_size/2.)*coefs_2[variable_name]
        for variable_name in variables.keys():
            coefs_3[variable_name] = self.variables[variable_name]['equation'](self,temp_variables,params)

        coefs_4 = {}
        for variable_name in variables.keys():
            temp_variables[variable_name] = variables[variable_name] + self.step_size*coefs_3[variable_name]
        for variable_name in variables.keys():
            coefs_4[variable_name] = self.variables[variable_name]['equation'](self,temp_variables,params)

        new_variables = {}
        for variable_name in variables.keys():
            new_variables[variable_name] = variables[variable_name] + (self.step_size/6.)*(coefs_1[variable_name]+2*coefs_2[variable_name]+2*coefs_3[variable_name]+coefs_4[variable_name])

        return new_variables





### BEGIN MainWindow class ###

class MainWindow(TemplateBaseClass,Modele):
    def __init__(self):

        # Create variables and parameters
        Modele.__init__(self)

        # Extra useful attributes
        self.fps             = None
        self.lastTime        = pgtime()
        self.colors_dict     = {'b':{'rgb':(31,119,180),'hex':'#1f77b4'},'o':{'rgb':(255,127,14),'hex':'#ff7f0e'},'g':{'rgb':(44,160,44),'hex':'#2ca02c'},'r':{'rgb':(214,39,40),'hex':'#d62728'},'p':{'rgb':(148,103,189),'hex':'#9467bd'},'y':{'rgb':(255,255,0),'hex':'#ffff00'}}
        self.flag_colormaps  = 1
        self.colormaps_list  = ['thermal','yellowy','greyclip','grey','viridis','inferno']

        # Load UI
        TemplateBaseClass.__init__(self)
        self.setWindowTitle('Graphical User Interface for Differential Equations (GUIDE)')

        # Create the main window
        self.ui = WindowTemplate()
        self.ui.setupUi(self)
        try: self.resize(*self.window_size)
        except: pass

        # Set main theme from self.window_params['theme']
        if 'theme' in self.__dict__.keys() and self.theme == 'dark':
            QtGui.QApplication.setStyle("Fusion")
            self.palette = self.palette()
            self.palette.setColor(QtGui.QPalette.Window, QtGui.QColor(53, 53, 53))
            self.palette.setColor(QtGui.QPalette.WindowText, QtCore.Qt.white)
            self.palette.setColor(QtGui.QPalette.Base, QtGui.QColor(25, 25, 25))
            self.palette.setColor(QtGui.QPalette.AlternateBase, QtGui.QColor(53, 53, 53))
            self.palette.setColor(QtGui.QPalette.ToolTipBase, QtCore.Qt.black)
            self.palette.setColor(QtGui.QPalette.ToolTipText, QtCore.Qt.white)
            self.palette.setColor(QtGui.QPalette.Text, QtCore.Qt.white)
            self.palette.setColor(QtGui.QPalette.Button, QtGui.QColor(53, 53, 53))
            self.palette.setColor(QtGui.QPalette.ButtonText, QtCore.Qt.white)
            self.palette.setColor(QtGui.QPalette.BrightText, QtCore.Qt.red)
            self.palette.setColor(QtGui.QPalette.Link, QtGui.QColor(42, 130, 218))
            self.palette.setColor(QtGui.QPalette.Highlight, QtGui.QColor(42, 130, 218))
            self.palette.setColor(QtGui.QPalette.HighlightedText, QtCore.Qt.black)
            self.setPalette(self.palette)

        # Button, sliders and spinboxes drawn in qtdesigner
        #ICs_button
        self.ui.ICs_button.clicked.connect(self.update_ICs_button)
        self.ui.ICs_button.keyPressEvent = self.keyPressEvent
        #nstep_slider
        self.ui.nstep_slider.setRange(1,int(self.array_size/10))
        self.ui.nstep_slider.setValue(self.nstep_update_plot)
        self.ui.nstep_slider.valueChanged.connect(self.update_nstep_slider)
        #nstep_spinbox
        self.ui.nstep_spinbox.setRange(1,int(self.array_size/10))
        self.ui.nstep_spinbox.setSingleStep(1)
        self.ui.nstep_spinbox.setValue(self.nstep_update_plot)
        self.ui.nstep_spinbox.setKeyboardTracking(False)  # emit signal only when enter is pressed
        self.ui.nstep_spinbox.valueChanged.connect(self.update_nstep_spinbox)
        #fps_label
        self.update_fps_label()
        #record_label
        self.ui.record_label.setText(' Rec. ')

        ########################## BEGIN figure layout and docks ##########################
        # Dock declaration and initial placement
        self.main_dock_area = self.ui.dock_area
        for dock_name in self.docks.keys():
            self.add_dock(dock_name) # add 'dock' and 'region' keywords into self.docks[dock_name]

        # Declaration of the plots in respective docks
        accepted_dock_types = ['plot1D','plot2D','image']
        assert self.docks[dock_name]['type'] in accepted_dock_types, f"Dock type '{self.docks[dock_name]['type']}' not understood. Dock type must be in {accepted_dock_types}"

        flag2 = 0
        alpha_factor_linearregion = 60  # 0 -> 255
        for dock_name in self.docks.keys():
            if self.docks[dock_name]['type'] == 'plot1D':
                self.create_PlotWidget(dock_name) # add 'actual_plot' keyword into self.docks[dock_name]

                # Attribution of the curves to the plots
                flag = 0
                self.docks[dock_name]['curve'] = {}
                for variable in self.variables.keys():
                    if 'dock' in self.variables[variable].keys():
                        if dock_name in self.variables[variable]['dock']:
                            self.docks[dock_name]['curve'][variable] = self.docks[dock_name]['actual_plot'].plot(pen=self.colors_dict[list(self.colors_dict.keys())[np.mod(flag,len(self.colors_dict))]]['rgb'])
                    else:
                        self.docks[dock_name]['curve'][variable] = self.docks[dock_name]['actual_plot'].plot(pen=self.colors_dict[list(self.colors_dict.keys())[np.mod(flag,len(self.colors_dict))]]['rgb'])
                    flag += 1

                if 'zoomOf' in self.docks[dock_name].keys():
                    relatedTo = self.docks[dock_name]['zoomOf']
                    # Create region and store in its according plot dict
                    self.docks[relatedTo]['region'][dock_name] = pg.LinearRegionItem([self.array_size/2.-self.array_size/30.,self.array_size/2.+self.array_size/30.],brush=self.colors_dict[list(self.colors_dict.keys())[np.mod(flag2,len(self.colors_dict))]]['rgb']+(alpha_factor_linearregion,))
                    self.docks[relatedTo]['region'][dock_name].setZValue(-10)
                    self.docks[relatedTo]['actual_plot'].addItem(self.docks[relatedTo]['region'][dock_name])
                    self.docks[relatedTo]['region'][dock_name].sigRegionChanged.connect(partial(self.update_zoom_plot,dock_name,relatedTo))
                    # Link region and zoom plot
                    self.docks[dock_name]['actual_plot'].sigXRangeChanged.connect(partial(self.update_xzoom_region,dock_name,relatedTo))
                    flag2 += 1

                    ### WARNING  Does not work probably due to an internal bug (waiting for answer)
                    #print('1',self.docks[dock_name]['actual_plot'].getViewBox().viewRange()[1])
                    #print('2',self.docks[relatedTo]['actual_plot'].getViewBox().viewRange()[1])
                    #self.docks[dock_name]['actual_plot'].setYLink(self.docks[relatedTo]['actual_plot'])
                    #print('1',self.docks[dock_name]['actual_plot'].getViewBox().viewRange()[1])
                    #print('2',self.docks[relatedTo]['actual_plot'].getViewBox().viewRange()[1])

                    self.update_zoom_plot(dock_name,relatedTo)


            elif self.docks[dock_name]['type'] == 'plot2D':
                self.create_PlotWidget(dock_name)

                # Attribution of the curves to the plots
                flag = 0
                self.docks[dock_name]['curve'] = {}
                for variable in self.variables.keys():
                    if 'dock' in self.variables[variable].keys():
                        # if element of 'dock' (variables/observables) is a dict
                        for element_variable_dock in self.variables[variable]['dock']:
                            if isinstance(element_variable_dock,dict):
                                if dock_name in element_variable_dock.keys():
                                    for real_dock_name in element_variable_dock.keys():
                                        # assert only two variables to plot
                                        assert len(element_variable_dock[real_dock_name]) == 2, f"list of variables/observables to plot on {real_dock_name} with dock type 'plot2D' must be exactly of length 2, provided was {len(element_variable_dock[real_dock_name])}"
                                        list_variables_to_plot = element_variable_dock[real_dock_name]
                                        # assert variables provided do exist
                                        for variables_to_plot in list_variables_to_plot:
                                            assert variables_to_plot in self.variables.keys(),f"variable '{variables_to_plot}' in 'dock' key of variable '{variable}' (variables/observables dictionnary) not understood. Must be in {list(self.variables.keys())}"
                                        self.docks[dock_name]['curve'][variable+'_plot2D_'+str(flag)] = {}
                                        self.docks[dock_name]['curve'][variable+'_plot2D_'+str(flag)]['curve'] = self.docks[dock_name]['actual_plot'].plot(pen=self.colors_dict[list(self.colors_dict.keys())[np.mod(flag,len(self.colors_dict))]]['rgb'])
                                        self.docks[dock_name]['curve'][variable+'_plot2D_'+str(flag)]['variables_to_plot'] = list_variables_to_plot
                                    flag += 1
                                else:
                                    print(f"WARNING: check validity of dock_names you provided in the variables/observable dictionnary: {element_variable_dock.keys()}'")
                if flag == 0:  # Nothing plotted on the 'plot2D'
                    print(f"WARNING: nothing has been plotted on the 'plot2D' dock with name '{dock_name}'")

                if 'zoomOf' in self.docks[dock_name].keys():
                    pass

            elif self.docks[dock_name]['type'] == 'image':
                self.create_ImageView(dock_name)
                self.docks[dock_name]['actual_plot'].keyPressEvent = self.keyPressEvent

        #self.docks[dock_name]['actual_plot'].enableAutoRange('xy', True)
        ########################## END figure layout and docks ##########################



        ############################ BEGIN Trees declaration ############################
        # Variables Tree
        self.tree = self.ui.tree
        self.tree.setColumnCount(3)
        self.tree.keyPressEvent = self.keyPressEvent # allow keys catching for focus on trees
        self.tree.setHeaderLabels(['Variables','IC','plot'])

        flag = 0
        for variable in self.variables.keys():
            temp = pg.TreeWidgetItem([variable])
            temp.setForeground(0,QtGui.QBrush(QtGui.QColor(self.colors_dict[list(self.colors_dict.keys())[np.mod(flag,len(self.colors_dict))]]['hex'])))

            # Create linedit (variables only)
            if not self.variables[variable]['observable']:
                self.variables[variable]['lineedit'] = QtGui.QLineEdit()
                temp.setWidget(1, self.variables[variable]['lineedit'])
                self.variables[variable]['lineedit'].setText(str(self.variables[variable]['value'][-1])) # set initial value
                self.variables[variable]['lineedit'].returnPressed.connect(partial(self.update_lineedit_variable,variable))

            # Create checkbox
            self.variables[variable]['checkbox'] = QtGui.QCheckBox()
            temp.setWidget(2, self.variables[variable]['checkbox'])
            self.tree.addTopLevelItem(temp)
            self.variables[variable]['checkbox'].setChecked(self.variables[variable]['plot']) # set initial state
            self.variables[variable]['checkbox'].keyPressEvent = self.keyPressEvent # connect keys
            self.variables[variable]['checkbox'].stateChanged.connect(partial(self.update_checkbox_variable,variable)) # connect checkbox
            flag += 1


        # Params Tree
        self.tree_params = self.ui.tree_params
        self.tree_params.setColumnCount(3)
        self.tree_params.keyPressEvent = self.keyPressEvent
        self.tree_params.setHeaderLabels(['Params','value','slider'])

        self.spinbox_precision = 3
        for param in self.params.keys():
            self.params[param]['slider_conversion_factor'] = int(1./self.params[param]['step'])   # To test was: 5000 *10000
            temp = pg.TreeWidgetItem([param])
            # Spin boxes
            self.params[param]['spinbox'] = QtGui.QDoubleSpinBox()
            self.params[param]['spinbox'].setRange(self.params[param]['min'],self.params[param]['max'])
            self.params[param]['spinbox'].setSingleStep(self.params[param]['step'])
            if isinstance(self.params[param]['step'],int):
                self.params[param]['spinbox'].setDecimals(0)
            else:
                self.params[param]['spinbox'].setDecimals(self.spinbox_precision)
            temp.setWidget(1, self.params[param]['spinbox'])
            self.tree_params.addTopLevelItem(temp)
            self.params[param]['spinbox'].setValue(self.params[param]['value'][-1])
            self.params[param]['spinbox'].setKeyboardTracking(False) # emit signal only when enter is pressed
            self.params[param]['spinbox'].valueChanged.connect(partial(self.update_slider_params,param))
            # Sliders
            self.params[param]['slider'] = QtGui.QSlider()
            self.params[param]['slider'].setRange(int(self.params[param]['min']*self.params[param]['slider_conversion_factor']),int(self.params[param]['max']*self.params[param]['slider_conversion_factor']))
            self.params[param]['slider'].setSingleStep(1)      # integers only
            self.params[param]['slider'].setOrientation(QtCore.Qt.Orientation.Horizontal)  # horizontale
            temp.setWidget(2, self.params[param]['slider'])
            self.tree.addTopLevelItem(temp)
            value = np.round(self.params[param]['value'][-1]*self.params[param]['slider_conversion_factor'],self.spinbox_precision) # convert in slider integer unit
            self.params[param]['slider'].setValue(int(value))
            self.params[param]['slider'].valueChanged.connect(partial(self.update_spinbox_params,param))


        # Kernel Tree
        self.tree_kernels = self.ui.tree_kernels
        self.tree_kernels.setColumnCount(2)
        self.tree_kernels.keyPressEvent = self.keyPressEvent
        self.tree_kernels.setHeaderLabels(['Kernels',''])

        # Create a group of buttons to allow "exclusive" behavior
        self.group_buttons_kernels = QtGui.QButtonGroup()
        self.group_buttons_kernels.setExclusive(True)
        for kernel in self.kernels.keys():
            self.kernels[kernel]['checkbox'] = QtGui.QCheckBox()
            self.group_buttons_kernels.addButton(self.kernels[kernel]['checkbox'], 1)
            temp = pg.TreeWidgetItem([kernel])
            temp.setWidget(1, self.kernels[kernel]['checkbox'])
            self.tree_kernels.addTopLevelItem(temp)
            if kernel == self.kernel:
                self.kernels[kernel]['checkbox'].setChecked(True) # set initial state
            self.kernels[kernel]['checkbox'].keyPressEvent = self.keyPressEvent

        self.group_buttons_kernels.buttonClicked.connect(self.update_checkbox_kernel)
        #############################  END Trees declaration  ############################

        # Start showing the window
        self.show()

        # Connect timer to update the figure
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.run_simulator)
        self.timer.start(10)

        # Initial window states
        if not self.streaming: self.timer.stop(); self.run_simulator()
        self.update_pause_indicator()
        self.update_record_state_indicator()

        # If starts recording from beginning
        if self.record_state:
            self.toggle_record_state()
            self.keyPressEvent("r")

        self.t = 0

    ################################ BEGIN plots update ###################################
    def update_zoom_plot(self,dock_name,relatedTo):
        self.docks[dock_name]['actual_plot'].setXRange(*self.docks[relatedTo]['region'][dock_name].getRegion(), padding=0)
    def update_xzoom_region(self,dock_name,relatedTo):
        #print('1',self.docks[dock_name]['actual_plot'].getViewBox().viewRange()[1])
        #print('2',self.docks[relatedTo]['actual_plot'].getViewBox().viewRange()[1])
        self.docks[relatedTo]['region'][dock_name].setRegion(self.docks[dock_name]['actual_plot'].getViewBox().viewRange()[0])

    def update_plots(self):
        for dock_name in self.docks.keys():
            if self.docks[dock_name]['type'] == 'plot1D':
                for variable in self.variables.keys():
                    if self.variables[variable]['plot']:
                        if 'dock' in self.variables[variable].keys():
                            if dock_name in self.variables[variable]['dock']:
                                self.docks[dock_name]['curve'][variable].setData(self.variables[variable]['value'])
                        else:
                            self.docks[dock_name]['curve'][variable].setData(self.variables[variable]['value'])
            elif self.docks[dock_name]['type'] == 'plot2D':
                # plot the variable names that are pre stored in dock dict
                for curve2D in self.docks[dock_name]['curve']:
                    # if variables specified, index 0 is to be plot
                    if self.variables[self.docks[dock_name]['curve'][curve2D]['variables_to_plot'][0]]['plot']:
                        self.docks[dock_name]['curve'][curve2D]['curve'].setData(self.variables[self.docks[dock_name]['curve'][curve2D]['variables_to_plot'][0]]['value'],self.variables[self.docks[dock_name]['curve'][curve2D]['variables_to_plot'][1]]['value'])
            elif self.docks[dock_name]['type'] == 'image':
                for variable in self.variables.keys():
                    if 'dock' in self.variables[variable].keys():
                        if self.variables[variable]['plot']:
                            if dock_name in self.variables[variable]['dock']:
                                self.docks[dock_name]['actual_plot'].setImage(self.variables[variable]['value'])
        # Update fps_label
        self.update_fps_label()

    def run_simulator(self):
        # Calculation
        for i in range(self.nstep_update_plot):
            self.simulator()

            # If recording
            if self.record_state and (self.nstep%self.nstep_record == 0):  # record every self.nstep_record
                self.append_to_dataframe()

            # Update main plots every self.nstep_update_plot (last occurence of the loop)
            if i==self.nstep_update_plot-1:
                self.update_plots()

            # Update time_stamp and parameter dict last (then saved correspond to calculation)
            self.time_stamp[:-1] = self.time_stamp[1:]
            self.time_stamp[-1] += self.step_size
            self.nstep          += 1

            for param in self.params.keys():
                self.params[param]['value'][:-1] = self.params[param]['value'][1:]

            # Fix app freezing on Windows systems  (if event occurs must process it)
            QtCore.QCoreApplication.processEvents()
            
    #################################  END plots update  ###################################


    def keyPressEvent(self, event):
        """ Set keyboard interactions """

        try: key = event.text()
        except: key = event  # allow calling keys programatically

        if key in list(self.user_defined_keyPressEvent.keys()):  # Interprete keys defined user file
            self.user_defined_keyPressEvent[key](self,{key:value['value'] for (key,value) in self.variables.items()},{key:value['value'][-1] for (key,value) in self.params.items()})
        elif key == ' ':
            self.toggle_streaming()
        elif key == 'q':
            sys.exit()
        elif key == 'h':
            previous_streaming_state = self.streaming
            if previous_streaming_state: self.toggle_streaming()
            self.display_help()
            if previous_streaming_state: self.toggle_streaming()
        elif key == 's' or key == 'r':
            previous_streaming_state = self.streaming
            if previous_streaming_state: self.toggle_streaming() # pause it
            if   key=='s': self.save() # query filename and save initial screenshot
            elif key=='r':
                if not self.record_state:
                    self.save(record=True)
                else:
                    self.toggle_record_state()
                    self.save_screenshot(self.filename_to_record_no_ext+'_END.png')
                    self.save_appended_dataframe()
                    self.filename_to_record_no_ext = None
            if previous_streaming_state: self.toggle_streaming()
        elif key == 'i':
            self.change_ICs_variable()
        elif key == 'c':
            self.update_images_colormap()
        else:
            if key != "" and event.key() != QtCore.Qt.Key_Return:
                print(f'Keyboard event "{key}" not None')


    def create_PlotWidget(self,dock_name):
        self.docks[dock_name]['actual_plot'] = pg.PlotWidget(**{key:value for key,value in self.docks[dock_name].items() if key not in ['dock','type','position','relativeTo','size','zoomOf','region']})
        self.docks[dock_name]['dock'].addWidget(self.docks[dock_name]['actual_plot'])

    def create_ImageView(self,dock_name):
        # Item for displaying image data
        pl = pg.PlotItem()  # to get axis
        img = pg.ImageItem(axisOrder='row-major')  # to rotate 90 degree

        # Create an ImageView Widget
        self.docks[dock_name]['actual_plot'] = pg.ImageView(view=pl,imageItem=img,**{key:value for key,value in self.docks[dock_name].items() if key not in ['dock','type','position','relativeTo','size','zoomOf','region']})
        # Set initial states
        self.docks[dock_name]['actual_plot'].view.invertY(False)
        self.docks[dock_name]['actual_plot'].view.setAspectLocked(False)
        self.docks[dock_name]['actual_plot'].view.disableAutoRange(True)
        self.docks[dock_name]['actual_plot'].ui.menuBtn.hide()
        #self.docks[dock_name]['actual_plot'].ui.menuBtn.show()
        #self.docks[dock_name]['actual_plot'].ui.histogram.hide()
        #self.docks[dock_name]['actual_plot'].ui.roiBtn.hide()

        # Set colormap to be used
        gradient = Gradients[self.colormaps_list[self.flag_colormaps]]
        cmap = pg.ColorMap(pos=[c[0] for c in gradient['ticks']],color=[c[1] for c in gradient['ticks']], mode=gradient['mode'])
        self.docks[dock_name]['actual_plot'].setColorMap(cmap)

        self.docks[dock_name]['dock'].addWidget(self.docks[dock_name]['actual_plot'])

    def add_dock(self,dock_name):
        ''' Add a dock to the main window '''
        if 'relativeTo' in self.docks[dock_name].keys():
            relativeto_dock_name = self.docks[dock_name]['relativeTo']
            assert 'dock' in self.docks[relativeto_dock_name].keys(), f"Dock '{relativeto_dock_name}' not understood. Docks that are 'relativeTo' another must be defined after it in the dictionnary of docks for consistent behavior"
        self.docks[dock_name]['region'] = {}  # 'region' key to be used later
        self.docks[dock_name]['dock'] = Dock(dock_name, size=self.docks[dock_name]['size'], closable=True)
        self.main_dock_area.addDock(**{key:value for key,value in self.docks[dock_name].items() if key in ['dock','position','relativeTo']})  # key used: 'dock', 'position' and 'relativeTo'

    def repaint_all_plots(self):
        for dock_name in self.docks.keys():
            if 'actual_plot' in self.docks[dock_name]:
                self.docks[dock_name]['actual_plot'].repaint()

    def toggle_streaming(self):
        self.streaming = not(self.streaming)
        self.update_pause_indicator()
    def update_pause_indicator(self):
        if self.streaming:
            self.ui.run_label.setStyleSheet("QLabel {border: 3px solid %s; background-color : %s; color : %s; }" %('#000000',self.colors_dict['g']['hex'],(0,0,0)))
            self.ui.run_label.setText('    Run    ')
            self.timer.start(10)
        else:
            self.ui.run_label.setStyleSheet("QLabel {border: 3px solid %s; background-color : %s; color : %s; }" %('#000000',self.colors_dict['r']['hex'],(0,0,0)))
            self.ui.run_label.setText('   Stop    ')
            self.timer.stop()
        self.ui.run_label.repaint()

    def update_images_colormap(self):
        self.flag_colormaps += 1
        cmap_name = self.colormaps_list[np.mod(self.flag_colormaps,len(self.colormaps_list))]
        gradient = Gradients[cmap_name]
        cmap     = pg.ColorMap(pos=[c[0] for c in gradient['ticks']],color=[c[1] for c in gradient['ticks']], mode=gradient['mode'])
        for dock_name in self.docks.keys():
            if self.docks[dock_name]['type'] == 'image':
                if 'actual_plot' in self.docks[dock_name]:
                    self.docks[dock_name]['actual_plot'].setColorMap(cmap)

        self.repaint_all_plots()

    def update_record_state_indicator(self):
        if self.record_state:
            self.ui.record_label.setStyleSheet("border: 3px solid %s; border-radius: 22px; background-color : %s; color : %s" %('#000000',self.colors_dict['r']['hex'],(0,0,0)))
        else:
            self.ui.record_label.setStyleSheet("border: 3px solid %s; border-radius: 22px; background-color : %s; color : %s" %('#000000','#000000','#000000'))
        self.ui.record_label.repaint()

    def update_ICs_button(self):
        for variable in self.variables.keys():
            if not self.variables[variable]['observable']:
                value = np.array(self.variables[variable]['init_cond']).astype(self.variables[variable]['type'])  # convert to array to be able to astype
                self.variables[variable]['lineedit'].setText(str(value)) # set initial value
            self.variables[variable]['value'] = self.variables[variable]['init_cond'] * np.ones(self.array_size).astype(self.variables[variable]['type'])

    def display_help(self):
        # Message must be a list of each line to display
        text_help_dialog  = ['Important Notes:','- (keyboard keys) do not work when focus is given to lineedits or spinboxes','- ("image" plots) you must pause to modify the aspect ratio, zoom or histogram range']
        text_help_dialog += ['']
        text_help_dialog += ['Usable keyboard keys:','- "  ":   toggle run/stop','- "q":   close the window','- "h":   display this help message','- "s":   save a snapshot and a dataframe','- "r":   toggle recording, save snapshots at start/end','- "i":   apply all variables ICs','- "c":   change the colormap to be use to draw "image" plots']
        text_help_dialog += ['']
        text_help_dialog += ['Defined variables and observables:']
        for variable in self.variables.keys():
            temp  = '- "'
            temp += variable+'"'
            if self.variables[variable]['observable']: temp += ' (observable)'
            elif not self.variables[variable]['observable']: temp += ' (variable)'
            if 'help' in self.variables[variable].keys(): temp += f":   {self.variables[variable]['help']}"
            text_help_dialog += [temp]
        text_help_dialog += ['']
        text_help_dialog += ['Defined parameters:']
        for param in self.params.keys():
            temp  = '- "'
            temp += param+'"'
            if 'help' in self.params[param].keys(): temp += f",  help:  {self.params[param]['help']}"
            for key in self.params[param].keys():
                if key in ['min','max','step','value']:
                    if key=='value':
                        temp += f",  {key}:  {self.params[param][key][-1]}"
                    else:
                        temp += f",  {key}:  {self.params[param][key]}"
            text_help_dialog += [temp]

        help_dialog = ScrollMessageBox(text_help_dialog,size_help=(850,600))
        help_dialog.setWindowTitle('Help message')
        help_dialog.exec_()


    ################################# BEGIN save ###################################
    def save(self,record=False,filename_to_save_no_ext=None):

        self.filename_to_save_no_ext = filename_to_save_no_ext
        if self.filename_to_save_no_ext is None:
            save_dialog = QtGui.QFileDialog()
            save_dialog.setFileMode(QtGui.QFileDialog.AnyFile)
            save_dialog.setFilter("Output files (*.png *.xlsx)")
            save_dialog.setWindowTitle("Saving files: screenshot, traces and window state")
            if save_dialog.exec_():
                filename_provided = save_dialog.selectedFiles()[0]
                if '.' in filename_provided:
                    self.filename_to_save_no_ext = filename_provided.rstrip('.'+filename_provided.split('.')[-1])  # remove extension if provided. WARNING if "." is written but no extension is provided
                else:
                    self.filename_to_save_no_ext = filename_provided

                # Build a dict of the existing conflicting files
                existing_filename_dict = {}
                for filename in [self.filename_to_save_no_ext+'.png',self.filename_to_save_no_ext+'.xlsx']:
                    if os.path.exists(filename):
                        existing_filename_dict[filename] = {}
                        existing_filename_dict[filename]['name'] = filename.split("/")[-1]
                        existing_filename_dict[filename]['path'] = filename.rstrip(filename.split("/")[-1])
                        existing_filename_dict[filename]['path']

                # Open a confirmation window if filename_provided exists
                if len(existing_filename_dict) > 0:
                    file_exists_dialog = QtGui.QMessageBox()
                    file_exists_dialog.setIcon(QtGui.QMessageBox.Warning)
                    file_exists_dialog.setWindowTitle('Warning: file already exists')
                    names = '" and "'.join([existing_filename_dict[key]['name'] for key in existing_filename_dict.keys()])
                    path = existing_filename_dict[list(existing_filename_dict.keys())[0]]['path']
                    if len(existing_filename_dict) > 1: extra_text = ['s','','them','them','their']
                    elif len(existing_filename_dict) == 1: extra_text = ['','s','it','it','its']
                    file_exists_dialog.setText(f'File{extra_text[0]} named "{names}" already exist{extra_text[1]} at location "{path}". Do you want to replace {extra_text[2]}?')
                    file_exists_dialog.setInformativeText(f'Replacing {extra_text[3]} will overwrite {extra_text[4]} contents forever.')
                    file_exists_dialog.setStandardButtons(QtGui.QMessageBox.Save|QtGui.QMessageBox.Cancel)
                    file_exists_dialog.setDefaultButton(QtGui.QMessageBox.Cancel)
                    file_exists_dialog.buttonClicked.connect(self.overwrite_buttons)
                    file_exists_dialog.exec_()

            save_dialog.close()

        # if closing the window or chose not to overwrite => no filename
        if self.filename_to_save_no_ext is None: return

        # save screenshot
        time.sleep(0.05) # wait for save_dialog to close before the snapshot
        add_text = '_START' if record else ''
        self.save_screenshot(self.filename_to_save_no_ext+f"{add_text}.png")
        # save dataframe with variables, observables and parameter values
        self.save_dataframe(self.filename_to_save_no_ext+'.xlsx')
        if record:
            self.list_to_record            = []
            self.filename_to_record_no_ext = self.filename_to_save_no_ext
            self.toggle_record_state()

    def overwrite_buttons(self,event):
        button_pressed = event.text()
        if button_pressed == 'Cancel':
            self.filename_to_save_no_ext = None
        elif button_pressed == 'Save':
            return
    def toggle_record_state(self):
        self.record_state = not(self.record_state)
        self.update_record_state_indicator()

    def save_screenshot(self,filename):
        """ Save a screenshot of the main_splitter (the whole "main" window) """
        screenshot = QtGui.QPixmap.grabWindow(self.ui.main_splitter.winId())
        screenshot.save(filename, 'png')
        print(f'File "{filename}" saved')
    def save_dataframe(self,filename):
        data_frame = self.build_dataframe_to_save()
        data_frame.to_excel(filename,index=False)
        print(f'File "{filename}" saved')
    def save_appended_dataframe(self,sheet_name='Sheet1'):
        writer = pd.ExcelWriter(self.filename_to_record_no_ext+'.xlsx', engine='openpyxl')
        writer.book = load_workbook(self.filename_to_record_no_ext+'.xlsx')
        startrow = writer.book[sheet_name].max_row
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        self.data_frame_to_record = pd.DataFrame(self.list_to_record)
        self.data_frame_to_record.to_excel(writer, sheet_name,header=False,index=False, startrow=startrow)
        writer.save()
        filename = self.filename_to_record_no_ext+'.xlsx'
        print(f'File "{filename}" appended')

    def build_dataframe_to_save(self):
        data_frame      = pd.DataFrame()
        data_frame['time'] = self.time_stamp
        for variable in self.variables.keys():
            if 'calculation_size' in self.variables[variable].keys() and self.variables[variable]['calculation_size']:
                continue
            data_frame[variable] = self.variables[variable]['value']
        for param in self.params.keys():
            data_frame[param] = self.params[param]['value']
        return data_frame

    def append_to_dataframe(self):
        list_to_data_frame = []
        list_to_data_frame.append(self.time_stamp[-1])
        for variable in self.variables.keys():
            if 'calculation_size' in self.variables[variable].keys() and self.variables[variable]['calculation_size']:
                continue
            list_to_data_frame.append(self.variables[variable]['value'][-1])
        for param in self.params.keys():
            list_to_data_frame.append(self.params[param]['value'][-1])
        self.list_to_record.append(list_to_data_frame)
    #################################  END save  ###################################

    def update_fps_label(self):
        self.time_now = pgtime()
        dt = self.time_now - self.lastTime
        self.lastTime = self.time_now
        if self.fps is None:
            self.fps = 1.0/dt
        else:
            s = np.clip(dt*3., 0, 1)
            self.fps = self.fps * (1-s) + (1.0/dt) * s
            self.ui.fps_label.setText('{:05.2f} fps'.format(self.fps))

    def update_checkbox_kernel(self):
        for kernel in self.kernels.keys():
            if self.kernels[kernel]['checkbox'].isChecked():
                self.kernel = kernel

    def update_checkbox_variable(self,variable):
        if self.variables[variable]['checkbox'].isChecked():
            self.variables[variable]['plot'] = True
        else:
            self.variables[variable]['plot'] = False
            # Somewhat duplicate of self.update_plots to clear only once
            for dock_name in self.docks.keys():
                if self.docks[dock_name]['type'] == 'plot1D':
                    if 'dock' in self.variables[variable].keys():
                        if dock_name in self.variables[variable]['dock']:
                            self.docks[dock_name]['curve'][variable].clear()
                    else:
                        self.docks[dock_name]['curve'][variable].clear()
                elif self.docks[dock_name]['type'] == 'plot2D':
                    if 'dock' in self.variables[variable].keys():
                        for curve2D in self.docks[dock_name]['curve']:
                            if variable == self.docks[dock_name]['curve'][curve2D]['variables_to_plot'][0]:
                                self.docks[dock_name]['curve'][curve2D]['curve'].clear()
                elif self.docks[dock_name]['type'] == 'image':
                    if 'dock' in self.variables[variable].keys():
                        if dock_name in self.variables[variable]['dock']:
                            self.docks[dock_name]['actual_plot'].clear()
                    else:
                        self.docks[dock_name]['actual_plot'].clear()

        # Force repainting all the 'actual_plot' (shouldn't be necessary => better being sure)
        self.update_plots()
        self.repaint_all_plots()

    def update_lineedit_variable(self,variable):
        """ Update the variables line edit """
        types = [complex,float,int]
        try:
            value = self.variables[variable]['lineedit'].text().replace(' ','')
            for typ in types:
                if isinstance(self.variables[variable]['value'][-1],typ):
                    self.variables[variable]['value'][-1] = typ(value)
        except ValueError:
            print(f'Input {value if len(value) else "None"} not a {typ.__name__} data type')


    def update_spinbox_params(self,param):
        value = np.round(self.params[param]['slider'].value()/self.params[param]['slider_conversion_factor'],self.spinbox_precision)
        if value <= self.params[param]['max'] and value >= self.params[param]['min']:
            self.params[param]['value'][-1] = value  # For simona
            self.params[param]['spinbox'].setValue(value)

    def update_slider_params(self,param):
        value = int(np.round(self.params[param]['spinbox'].value()*self.params[param]['slider_conversion_factor']))
        if isinstance(self.params[param]['step'],int):
            value = int(value)
        if value <= self.params[param]['max']*self.params[param]['slider_conversion_factor'] and value >= self.params[param]['min']*self.params[param]['slider_conversion_factor']:
            self.params[param]['value'][-1] = value/self.params[param]['slider_conversion_factor']
            self.params[param]['slider'].setValue(value)
        # Update observables and plots (works also for spinbox here as setValue calls this func.)
        self.update_observables()
        self.update_plots()

    def update_nstep_slider(self):
        value = self.ui.nstep_slider.value()
        self.nstep_update_plot = value
        self.ui.nstep_spinbox.setValue(value)
    def update_nstep_spinbox(self):
        value = self.ui.nstep_spinbox.value()
        self.nstep_update_plot = value
        self.ui.nstep_slider.setValue(value)
    def change_ICs_variable(self):
        for variable in self.variables.keys():
            if not self.variables[variable]['observable']:
                self.update_lineedit_variable(variable)


# Convenience Class for message with scroll bar
class ScrollMessageBox(QtGui.QMessageBox):
   def __init__(self, message_list, size_help=(850,600), *args, **kwargs):
      QtGui.QMessageBox.__init__(self, *args, **kwargs)
      scroll = QtGui.QScrollArea(self)
      scroll.setWidgetResizable(True)
      self.content = QtGui.QWidget()
      scroll.setWidget(self.content)
      lay = QtGui.QVBoxLayout(self.content)
      for item in message_list:
         lay.addWidget(QtGui.QLabel(item, self))
      self.layout().addWidget(scroll, 0, 0, 1, self.layout().columnCount())
      self.setStyleSheet("QScrollArea{min-width:%d px; min-height: %dpx}" %(size_help[0],size_help[1]))




## Start Qt event loop unless running in interactive mode or using pyside.
if __name__ == '__main__':
    ### BEGIN Start the window ###
    win = MainWindow()

    if (sys.flags.interactive != 1) or not hasattr(QtCore, 'PYQT_VERSION'):
        QtGui.QApplication.instance().exec_()
